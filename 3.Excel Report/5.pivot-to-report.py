from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

# Config
month = input("Enter Month Name: ")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "pivot_table.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, f"report_{month}.xlsx")

# Load workbook
wb = load_workbook(INPUT_FILE)
sheet = wb['Report']

# Active rows and columns
min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# Bar chart
# 
barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

barchart.title = 'Sales by Product line'
barchart.style = 5

sheet.add_chart(barchart, "B12")

# 
# Totals row (formulas)
# 
for i in range(min_column + 1, max_column + 1):
    col_letter = get_column_letter(i)
    cell = f'{col_letter}{max_row + 1}'
    sheet[cell] = f'=SUM({col_letter}{min_row + 1}:{col_letter}{max_row})'
    sheet[cell].style = 'Currency'

# 
# Header formatting
# 
sheet['A1'] = 'Sales Report'
sheet['A2'] = month.capitalize()
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

# 
# Save output
# 
wb.save(OUTPUT_FILE)

print("Report generated successfully.")
